"""
DataReader — 结构化数据读取器
==============================
将 CSV/Excel 文件解析为 StructuredData，包含列统计和完整数据行。

与 file_tools.parse_file() 的区别：
- parse_file() 返回管道分隔的文本表格（供 LLM 阅读）
- DataReader 返回结构化的 {col: value} 字典列表（供代码直接计算）

设计：纯 Python 实现，不依赖 pandas。使用 statistics 模块做数值计算。
"""

import csv
import io
import logging
from typing import Dict, Any, List, Optional
from pathlib import Path
from statistics import median, stdev, quantiles

logger = logging.getLogger(__name__)

from src.models.schemas import StructuredData, ColumnStats

logger = logging.getLogger(__name__)

ENCODINGS = ["utf-8-sig", "utf-8", "gbk", "gb2312", "utf-16", "cp1252", "latin-1"]
MAX_ROWS = 10000  # 超过此行数的文件截断


class DataReader:
    """结构化数据读取器 — 单例"""

    def read_file(self, filepath: str, max_rows: int = MAX_ROWS) -> List[StructuredData]:
        """读取 CSV/Excel 文件，返回 StructuredData 列表（Excel 每个 sheet 一个）"""
        path = Path(filepath)
        if not path.exists():
            raise FileNotFoundError(f"文件不存在: {filepath}")

        suffix = path.suffix.lower()
        if suffix == ".csv":
            return [self._read_csv(filepath, max_rows)]
        elif suffix in (".xlsx", ".xls"):
            return self._read_xlsx(filepath, max_rows)
        else:
            raise ValueError(f"不支持的文件格式: {suffix}，仅支持 CSV/Excel")

    def _read_csv(self, filepath: str, max_rows: int) -> StructuredData:
        """解析 CSV 为结构化数据"""
        path = Path(filepath)
        raw_bytes = path.read_bytes()

        # 编码检测
        content = None
        used_encoding = "utf-8"
        for enc in ENCODINGS:
            try:
                content = raw_bytes.decode(enc)
                used_encoding = enc
                break
            except (UnicodeDecodeError, LookupError):
                continue

        if content is None:
            raise ValueError(f"无法解码 CSV 文件: {filepath}")

        if used_encoding == "latin-1":
            logger.warning(
                f"CSV 文件 {path.name} 使用 latin-1 回退编码，"
                f"数据可能已乱码。建议转换为 UTF-8 格式。"
            )

        # 解析 CSV（流式：仅读取 max_rows + 1 行检测截断）
        reader = csv.reader(io.StringIO(content))
        try:
            first_row = next(reader)
        except StopIteration:
            return StructuredData(filename=path.name, row_count=0, col_count=0)

        # 第一行作为表头
        headers = [h.strip() for h in first_row]
        data_rows = []
        for i, row in enumerate(reader):
            if i >= max_rows:
                truncated = True
                break
            data_rows.append(row)
        else:
            truncated = False

        rows = []
        for row in data_rows:
            row_dict = {}
            for i, header in enumerate(headers):
                row_dict[header] = row[i].strip() if i < len(row) else ""
            rows.append(row_dict)

        return self._build_structured_data(path.name, headers, rows, None, used_encoding, truncated)

    def _read_xlsx(self, filepath: str, max_rows: int) -> List[StructuredData]:
        """解析 Excel 为结构化数据列表"""
        import openpyxl

        path = Path(filepath)
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        try:
            results = []

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                all_rows = list(ws.iter_rows(values_only=True))

                if not all_rows:
                    results.append(StructuredData(
                        filename=path.name, sheet_name=sheet_name, row_count=0, col_count=0
                    ))
                    continue

                # 第一行作为表头
                headers = [str(h).strip() if h is not None else f"Column_{i}"
                           for i, h in enumerate(all_rows[0])]
                data_rows = all_rows[1:]

                # 过滤全空行
                data_rows = [r for r in data_rows if any(c is not None for c in r)]

                truncated = len(data_rows) > max_rows
                if truncated:
                    data_rows = data_rows[:max_rows]

                rows = []
                for row in data_rows:
                    row_dict = {}
                    for i, header in enumerate(headers):
                        val = row[i] if i < len(row) else ""
                        row_dict[header] = str(val).strip() if val is not None else ""
                    rows.append(row_dict)

                results.append(self._build_structured_data(
                    path.name, headers, rows, sheet_name, "openpyxl", truncated
                ))

            return results if results else [StructuredData(filename=path.name, row_count=0, col_count=0)]
        finally:
            wb.close()

    def _build_structured_data(
        self, filename: str, headers: List[str], rows: List[Dict[str, Any]],
        sheet_name: Optional[str], used_encoding: str, truncated: bool
    ) -> StructuredData:
        """从 headers 和 rows 构建 StructuredData，包含列统计"""
        column_stats = {}
        for col in headers:
            values = [row.get(col, "") for row in rows]
            stats = self._compute_column_stats(col, values)
            column_stats[col] = stats

        return StructuredData(
            filename=filename,
            sheet_name=sheet_name,
            columns=headers,
            column_stats=column_stats,
            rows=rows,
            row_count=len(rows),
            col_count=len(headers),
            truncated=truncated,
            file_encoding=used_encoding,
        )

    def _compute_column_stats(self, name: str, values: List[Any]) -> ColumnStats:
        """计算单列的统计摘要"""
        count = len(values)
        non_null = [v for v in values if v is not None and str(v).strip() != ""]
        null_count = count - len(non_null)
        unique_count = len(set(str(v) for v in non_null))

        # 类型推断
        dtype, numeric_vals = self._infer_dtype(non_null)

        stats = ColumnStats(
            name=name, dtype=dtype, count=count,
            null_count=null_count, unique_count=unique_count,
        )

        if dtype == "numeric" and len(numeric_vals) >= 2:
            try:
                stats.min = min(numeric_vals)
                stats.max = max(numeric_vals)
                stats.mean = sum(numeric_vals) / len(numeric_vals)
                stats.median = median(numeric_vals)
                if len(numeric_vals) >= 3:
                    stats.std = stdev(numeric_vals)
                if len(numeric_vals) >= 4:
                    qs = quantiles(numeric_vals, n=4)
                    stats.q1 = qs[0]
                    stats.q3 = qs[2]
            except Exception as e:
                logger.debug("Column stats computation failed for '%s': %s", col_name, e)

        # 前5个样本
        stats.sample_values = [str(v) for v in values[:5]]

        return stats

    def _infer_dtype(self, values: List[Any]) -> tuple:
        """推断列的数据类型。>=80% 可转 float → numeric。"""
        numeric_count = 0
        numeric_vals = []
        for v in values:
            try:
                fv = float(v)
                numeric_count += 1
                numeric_vals.append(fv)
            except (ValueError, TypeError):
                pass

        if len(values) > 0 and numeric_count / len(values) >= 0.8:
            return "numeric", numeric_vals
        return "string", []

    def get_numeric_columns(self, data: StructuredData) -> List[str]:
        """返回所有数值列的名称"""
        return [c for c in data.columns if data.column_stats.get(c, ColumnStats(name=c)).dtype == "numeric"]

    def get_string_columns(self, data: StructuredData) -> List[str]:
        """返回所有字符串列的名称"""
        return [c for c in data.columns if data.column_stats.get(c, ColumnStats(name=c)).dtype != "numeric"]

    def get_column_values(self, data: StructuredData, col: str, typed: bool = True) -> List[Any]:
        """提取单列的所有值。typed=True 时数值列返回 float（跳过无法解析的值）。"""
        if col not in data.columns:
            return []
        stats = data.column_stats.get(col, ColumnStats(name=col))
        values = [row.get(col, "") for row in data.rows]
        if typed and stats.dtype == "numeric":
            result = []
            for v in values:
                try:
                    result.append(float(v))
                except (ValueError, TypeError):
                    continue  # 跳过非数值单元格（如空值、表头、脚注等）
            return result
        return [str(v) for v in values]


# 单例
_data_reader_instance: Optional[DataReader] = None


def get_data_reader() -> DataReader:
    global _data_reader_instance
    if _data_reader_instance is None:
        _data_reader_instance = DataReader()
    return _data_reader_instance
